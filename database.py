import sqlite3
from pathlib import Path
from datetime import datetime
import os


class Database:
    def __init__(self):
        # ✅ 智能路径处理（自动转义#号/空格/中文）
        app_data = Path.home() / "AppData" / "Roaming" / "DailyReport"
        app_data.mkdir(parents=True, exist_ok=True)
        self.db_path = app_data / "dailydata.db"
        self.init_db()
        print(f"✅ 数据库就绪 | 路径: {self.db_path}")

    def init_db(self):
        """初始化数据库表"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        # 任务表
        cursor.execute('''
                       CREATE TABLE IF NOT EXISTS daily_reports
                       (
                           id
                           INTEGER
                           PRIMARY
                           KEY
                           AUTOINCREMENT,
                           report_date
                           TEXT
                           NOT
                           NULL,
                           project
                           TEXT,
                           task_content
                           TEXT
                           NOT
                           NULL,
                           hours
                           REAL,
                           status
                           TEXT
                           DEFAULT
                           '进行中',
                           notes
                           TEXT,
                           created_time
                           DATETIME
                           DEFAULT
                           CURRENT_TIMESTAMP
                       )
                       ''')
        # 创建索引提升查询速度
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_date ON daily_reports(report_date)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_status ON daily_reports(status)')
        conn.commit()
        conn.close()

    # ========== CRUD 操作 ==========
    def save_report(self, date, project, task, hours, status, notes):
        """保存新日报"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO daily_reports (report_date, project, task_content, hours, status, notes) VALUES (?, ?, ?, ?, ?, ?)",
                (date, project, task, hours, status, notes)
            )
            conn.commit()
            report_id = cursor.lastrowid
            conn.close()
            return True, report_id
        except Exception as e:
            return False, str(e)

    def get_all_reports(self, limit=100):
        """获取所有日报（按时间倒序）"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""
                       SELECT id,
                              report_date,
                              project,
                              task_content,
                              hours,
                              status,
                              notes,
                              created_time
                       FROM daily_reports
                       ORDER BY created_time DESC LIMIT ?
                       """, (limit,))
        rows = cursor.fetchall()
        conn.close()
        return rows

    def get_reports_by_date(self, start_date, end_date):
        """按日期范围查询"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""
                       SELECT id,
                              report_date,
                              project,
                              task_content,
                              hours,
                              status,
                              notes,
                              created_time
                       FROM daily_reports
                       WHERE report_date BETWEEN ? AND ?
                       ORDER BY report_date DESC
                       """, (start_date, end_date))
        rows = cursor.fetchall()
        conn.close()
        return rows

    def update_report(self, report_id, date, project, task, hours, status, notes):
        """更新日报"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                           UPDATE daily_reports
                           SET report_date=?,
                               project=?,
                               task_content=?,
                               hours=?,
                               status=?,
                               notes=?
                           WHERE id = ?
                           """, (date, project, task, hours, status, notes, report_id))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            return False, str(e)

    def delete_report(self, report_id):
        """删除日报"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM daily_reports WHERE id=?", (report_id,))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            return False, str(e)

    def export_to_excel(self, rows, output_path):
        """导出到Excel（含格式）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "工作日报"

            # 表头
            headers = ["ID", "日期", "项目", "任务内容", "耗时(小时)", "状态", "备注", "创建时间"]
            ws.append(headers)

            # 样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # 数据
            for row in rows:
                # 转换时间格式
                created_time = row[7]
                if created_time:
                    try:
                        created_time = datetime.strptime(created_time, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M")
                    except:
                        pass
                ws.append([
                    row[0], row[1], row[2], row[3],
                    f"{row[4]:.1f}" if row[4] else "",
                    row[5], row[6], created_time
                ])

            # 列宽优化
            col_widths = [6, 12, 15, 40, 10, 10, 20, 18]
            for i, width in enumerate(col_widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            wb.save(output_path)
            return True, None
        except Exception as e:
            return False, str(e)